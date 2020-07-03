import openpyxl


class UpdateSettings:

    def __init__(self,
                 request_data_type,
                 request_data_dimensions,
                 request_data_sensitivity,
                 request_channel,
                 channel_pref,
                 channel_pref_num,
                 request_mode,
                 mode_pref,
                 mode_pref_num
                 ):

        wb = openpyxl.load_workbook("settings.xlsx")
        ws = wb["mode_network"]

        request_data_type = float(request_data_type)
        request_data_dimensions = float(request_data_dimensions)
        request_data_sensitivity = float(request_data_sensitivity)
        request_channel = float(request_channel)
        if channel_pref != "":
            channel_pref = float(channel_pref)
        channel_pref_num = float(channel_pref_num)
        request_mode = float(request_mode)
        if mode_pref != "":
            mode_pref = float(mode_pref)
        mode_pref_num = float(mode_pref_num)

        # Updating mode rec
        if request_data_type == 0:
            ws.cell(row=3, column=2).value = 0.5
            ws.cell(row=3, column=3).value = 0
            ws.cell(row=3, column=4).value = 0.5
        elif request_data_dimensions == 1:
            ws.cell(row=3, column=2).value = 0.5
            ws.cell(row=3, column=3).value = 0.2
            ws.cell(row=3, column=4).value = 0.3
        elif request_data_dimensions == 2:
            ws.cell(row=3, column=2).value = 0.2
            ws.cell(row=3, column=3).value = 0.7
            ws.cell(row=3, column=4).value = 0.1
        elif request_data_dimensions > 2:
            ws.cell(row=3, column=2).value = 0.1
            ws.cell(row=3, column=3).value = 0.9
            ws.cell(row=3, column=4).value = 0

        # Updating request mode
        ws.cell(row=2, column=2).value = 0
        ws.cell(row=2, column=3).value = 0
        ws.cell(row=2, column=4).value = 0
        ws.cell(row=2, column=request_mode + 2).value = 1

        # Updating mode pref
        if mode_pref != "":
            ws.cell(row=4, column=2).value = (1 - (mode_pref_num * 0.4 / 3 + 0.5)) / 2
            ws.cell(row=4, column=3).value = (1 - (mode_pref_num * 0.4 / 3 + 0.5)) / 2
            ws.cell(row=4, column=4).value = (1 - (mode_pref_num * 0.4 / 3 + 0.5)) / 2
            ws.cell(row=4, column=mode_pref + 2).value = mode_pref_num * 0.4 / 3 + 0.5
        else:
            ws.cell(row=4, column=2).value = 1/3
            ws.cell(row=4, column=3).value = 1/3
            ws.cell(row=4, column=4).value = 1/3

        ws = wb["channel_network"]

        # Updating channel rec
        ws.cell(row=4, column=3).value = request_data_sensitivity * 0.4 / 3 + 0.5
        ws.cell(row=4, column=2).value = 1 - (request_data_sensitivity * 0.4 / 3 + 0.5)

        # Updating request channel
        ws.cell(row=2, column=2).value = 0
        ws.cell(row=2, column=3).value = 0
        ws.cell(row=2, column=request_channel + 2).value = 1

        # Updating channel pref
        if channel_pref != "":
            ws.cell(row=3, column=2).value = 1 - (channel_pref_num * .5 / 3 + .5)
            ws.cell(row=3, column=3).value = 1 - (channel_pref_num * .5 / 3 + .5)
            ws.cell(row=3, column=channel_pref + 2).value = channel_pref_num * .5 / 3 + 0.5
        else:
            ws.cell(row=3, column=2).value = 1 / 2
            ws.cell(row=3, column=3).value = 1 / 2

        wb.save("settings.xlsx")